from openpyxl import load_workbook
from class_element_for_tree import BinaryTreeElement
from tkinter import *


def place_the_item_in_a_tree(element: BinaryTreeElement, position_x: int, position_y: int) -> None:
    if element.return_best_main_premise() is not None and element.find_best_entropy() is not None:
        print(element.return_conclusion_tab())
        element.calculate_conclusion_entropy()
        element.return_entropy_tab()
        element.return_information_gain()
        print(f"Max(I-Ej): {element.return_best_main_premise()} {element.find_best_entropy()}")
        element.draw_element_on_tree(position_x=position_x, position_y=position_y)
        # print(element.return_left_branch())
        # print(element.return_right_branch())
        print(element)


def create_an_option_to_table(sheet, min_row: int, min_col: int, max_row: int, max_col: int, table=None):
    if table is None:
        table = []
    for col in sheet.iter_cols(min_row=min_row, min_col=min_col, max_row=max_row, max_col=max_col):
        for cell in col:
            if cell.value is not None:
                table.append(cell.value)
    return table


def run_example():
    wb = load_workbook("Test.xlsx")

    sheet = wb.active

    premises = {}
    one_hundred_examples = {}

    option_of_premises = create_an_option_to_table(sheet, 2, 2, 16, 2)
    option_of_conclusion = create_an_option_to_table(sheet, 18, 1, 20, 1)

    for value in range(100):
        one_hundred_examples[value] = {}

    for col in sheet.iter_cols(min_row=2, min_col=1, max_row=14, max_col=1):
        for cell in col:
            if cell.value is not None:
                premises[cell.value] = ""

    for number, col in enumerate(sheet.iter_cols(min_row=2, min_col=3, max_row=20, max_col=102)):
        for index, cell in enumerate(col):
            if cell.value is not None:
                if 0 <= index <= 3:
                    premises["wiek"] = option_of_premises[index]
                elif 4 <= index <= 5:
                    premises["płeć"] = option_of_premises[index]
                elif 6 <= index <= 8:
                    premises["zar. mies."] = option_of_premises[index]
                elif 9 <= index <= 11:
                    premises["ost. posiłek"] = option_of_premises[index]
                elif 12 <= index <= 14:
                    premises["os. spoż. pizze"] = option_of_premises[index]
                elif 16 <= index <= 18:
                    premises["konkluzja(wybór pizzy):"] = option_of_conclusion[index - 16]
                    one_hundred_examples[number].update(premises)

    # for key, value in one_hundred_examples.items():
    #     print(f"{key} : {value}")

    window = Tk()
    window.geometry('1600x900')
    window.title("Drzewo decyzyjne")
    conclusion = "konkluzja(wybór pizzy):"
    main_premises = list(premises.keys())


    """"" Binary tree: N - node, R - right, L - left (L and R means the direction of movement in the tree) """

    """ LEVEL 0 """

    N = BinaryTreeElement(one_hundred_examples, conclusion, main_premises, option_of_premises,
                                    option_of_conclusion, window)
    place_the_item_in_a_tree(N, position_x=500, position_y=0)

    """ LEVEL 1 """

    L = BinaryTreeElement(N.return_left_branch(), conclusion, main_premises, option_of_premises,
                                    option_of_conclusion, window)
    place_the_item_in_a_tree(L, position_x=430, position_y=50)

    R = BinaryTreeElement(N.return_right_branch(), conclusion, main_premises, option_of_premises,
                                    option_of_conclusion, window)

    place_the_item_in_a_tree(R, position_x=700, position_y=50)

    """ LEVEL 2 """

    LL = BinaryTreeElement(L.return_left_branch(), conclusion, main_premises, option_of_premises,
                           option_of_conclusion, window)
    place_the_item_in_a_tree(LL, position_x=250, position_y=100)

    LR = BinaryTreeElement(L.return_right_branch(), conclusion, main_premises, option_of_premises,
                           option_of_conclusion, window)
    place_the_item_in_a_tree(LR, position_x=550, position_y=100)

    RL = BinaryTreeElement(R.return_left_branch(), conclusion, main_premises, option_of_premises,
                                    option_of_conclusion, window)
    place_the_item_in_a_tree(RL, position_x=650, position_y=100)

    RR = BinaryTreeElement(R.return_right_branch(), conclusion, main_premises, option_of_premises,
                           option_of_conclusion, window)
    place_the_item_in_a_tree(RR, position_x=850, position_y=100)

    """ LEVEL 3 """

    LLL = BinaryTreeElement(LL.return_left_branch(), conclusion, main_premises, option_of_premises,
                            option_of_conclusion, window)
    place_the_item_in_a_tree(LLL, position_x=200, position_y=150)

    LLR = BinaryTreeElement(LL.return_right_branch(), conclusion, main_premises, option_of_premises,
                            option_of_conclusion, window)
    place_the_item_in_a_tree(LLR, position_x=430, position_y=150)

    RRL = BinaryTreeElement(RR.return_left_branch(), conclusion, main_premises, option_of_premises,
                                    option_of_conclusion, window)
    place_the_item_in_a_tree(RRL, position_x=740, position_y=150)

    RRR = BinaryTreeElement(RR.return_right_branch(), conclusion, main_premises, option_of_premises,
                                    option_of_conclusion, window)
    place_the_item_in_a_tree(RRR, position_x=1000, position_y=150)

    """ LEVEL 4 """

    LLRL = BinaryTreeElement(LLR.return_left_branch(), conclusion, main_premises, option_of_premises,
                             option_of_conclusion, window)
    place_the_item_in_a_tree(LLRL, position_x=360, position_y=200)

    LLRR = BinaryTreeElement(LLR.return_right_branch(), conclusion, main_premises, option_of_premises,
                             option_of_conclusion, window)
    place_the_item_in_a_tree(LLRR, position_x=520, position_y=200)

    RRLL = BinaryTreeElement(RRL.return_left_branch(), conclusion, main_premises, option_of_premises,
                                    option_of_conclusion, window)
    place_the_item_in_a_tree(RRLL, position_x=660, position_y=200)

    RRLR = BinaryTreeElement(RRL.return_right_branch(), conclusion, main_premises, option_of_premises,
                                    option_of_conclusion, window)
    place_the_item_in_a_tree(RRLR, position_x=830, position_y=200)

    """ LEVEL 5 """

    LLRLL = BinaryTreeElement(LLRL.return_left_branch(), conclusion, main_premises, option_of_premises,
                              option_of_conclusion, window)
    place_the_item_in_a_tree(LLRLL, position_x=300, position_y=250)

    LLRLR = BinaryTreeElement(LLRL.return_right_branch(), conclusion, main_premises, option_of_premises,
                              option_of_conclusion, window)
    place_the_item_in_a_tree(LLRLR, position_x=420, position_y=250)

    RRLRL = BinaryTreeElement(RRLR.return_left_branch(), conclusion, main_premises, option_of_premises,
                                    option_of_conclusion, window)
    place_the_item_in_a_tree(RRLRL, position_x=760, position_y=250)

    RRLRR = BinaryTreeElement(RRLR.return_right_branch(), conclusion, main_premises, option_of_premises,
                              option_of_conclusion, window)
    place_the_item_in_a_tree(RRLRR, position_x=900, position_y=250)

    """ LEVEL 6 """

    LLRLRL = BinaryTreeElement(LLRLR.return_left_branch(), conclusion, main_premises, option_of_premises,
                               option_of_conclusion, window)
    place_the_item_in_a_tree(LLRLRL, position_x=350, position_y=300)

    LLRLRR = BinaryTreeElement(LLRLR.return_right_branch(), conclusion, main_premises, option_of_premises,
                               option_of_conclusion, window)
    place_the_item_in_a_tree(LLRLRR, position_x=600, position_y=300)

    RRLRRL = BinaryTreeElement(RRLRR.return_left_branch(), conclusion, main_premises, option_of_premises,
                                    option_of_conclusion, window)
    place_the_item_in_a_tree(RRLRRL, position_x=820, position_y=300)

    RRLRRR = BinaryTreeElement(RRLRR.return_right_branch(), conclusion, main_premises, option_of_premises,
                                    option_of_conclusion, window)
    place_the_item_in_a_tree(RRLRRR, position_x=1080, position_y=300)

    """ Legend for the binary tree"""

    salary = Label(window, text="zar. mies. - zarobki miesięczne danej osoby",
                   font=("Times New Roman", 20))

    meals = Label(window, text="ost. posiłek - czas od ostatniego posiłku danej osoby",
                  font=("Times New Roman", 20))

    persons_eating = Label(window, text="os. spoż. pizze - liczba osób spożywających pizzę",
                           font=("Times New Roman", 20))

    salary.place(x=50, y=500)
    meals.place(x=50, y=540)
    persons_eating.place(x=50, y=580)

    window.mainloop()

if __name__ == "__main__":
    run_example()
